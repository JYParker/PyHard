import re
import os
import zipfile
from openpyxl import load_workbook

patterns = {
    "주민등록번호": r"\b\d{6}-\d{7}\b",
    "전화번호": r"\b01[016789]-\d{3,4}-\d{4}\b",
    "이메일": r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-z]{2,}\b"
}

def sensitive_info(text):
    return any(re.search(pattern, text) for pattern in patterns)

def inspect_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
        return sensitive_info(content)
    
def inspect_xlsx(file_path):
    wb = load_workbook(file_path, data_only=True)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and sensitive_info(str(cell)):
                    return True
    return False

def inspect_zip(file_path):
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        zip_ref.extractall("temp_zip")
        for name in zip_ref.namelist():
            extension = os.path.splitext(name)[1].lower()
            full_path = os.path.join("temp_zip", name)
            if extension == '.txt' and inspect_txt(full_path):
                return True
            elif extension == '.xlsx' and inspect_xlsx(full_path):
                return True
    return False

def scan_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    result = False

    if ext == '.txt':
        result = inspect_txt(file_path)
    elif ext == '.xlsx':
        result = inspect_xlsx(file_path)
    elif ext == '.zip':
        result = inspect_zip(file_path)
    else:
        print(f"지원하지 않는 파일 형식")
        return False
    
    if result:
        print("민감정보 포함")
    else:
        print("민감정보가 없습니다.")
    return result

